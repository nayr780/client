#!/usr/bin/env python3
# queue_client_front.py — Backend WS bridge + REST, com status e logs em tempo real

import os
import time
import uuid
import json
import pickle
import asyncio
from typing import Dict, Any, List, Optional, Union, Tuple, Set

import requests
from fastapi import (
    FastAPI,
    HTTPException,
    UploadFile,
    File,
    Request,
    WebSocket,
    WebSocketDisconnect,
    Form,
    APIRouter,
)
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

# ===================== DEBUG HELPERS =====================

DEBUG = os.getenv("DEBUG_QUEUE_FRONT", "1").lower() in ("1", "true", "yes")


def dbg(*args):
    if DEBUG:
        ts = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        print(f"[dbg {ts}]", *args)


# ===================== CLIENTE WS RESP (inline) =====================

import websockets


class RedisRESPError(Exception):
    pass


class AsyncRedisWS:
    """
    Cliente mínimo RESP sobre WebSocket.
    Suporta: + - : $ * (simple string, error, integer, bulk, array)
    """

    def __init__(self, url: str, token: str = "", connect_timeout: float = 10.0, name: str = "cmd"):
        self.url = f"{url}?token={token}" if token else url
        self._ws: Optional[websockets.WebSocketClientProtocol] = None
        self._lock = asyncio.Lock()
        self._connect_timeout = connect_timeout
        self._name = name

    async def connect(self):
        if self._ws is not None and not getattr(self._ws, "closed", False):
            return
        dbg(f"[redis-{self._name}] connecting -> {self.url}")
        self._ws = await websockets.connect(self.url, max_size=None)
        dbg(f"[redis-{self._name}] connected")

    async def aclose(self):
        if self._ws is not None:
            try:
                await self._ws.close()
                dbg(f"[redis-{self._name}] closed")
            except Exception as e:
                dbg(f"[redis-{self._name}] close error:", repr(e))
            self._ws = None

    def _encode_cmd(self, *parts: Union[str, bytes, int, float]) -> bytes:
        out = bytearray()
        out.extend(f"*{len(parts)}\r\n".encode())
        for p in parts:
            if isinstance(p, (bytes, bytearray)):
                s = bytes(p)
            else:
                s = str(p).encode()
            out.extend(f"${len(s)}\r\n".encode())
            out.extend(s)
            out.extend(b"\r\n")
        return bytes(out)

    def _read_line(self, buf: bytearray, start: int = 0) -> Tuple[bytes, int]:
        i = buf.find(b"\r\n", start)
        if i == -1:
            raise RedisRESPError("incomplete")
        return bytes(buf[start:i]), i + 2

    def _parse_resp(self, data: bytes) -> Any:
        buf = bytearray(data)
        pos = 0

        def parse_at():
            nonlocal pos
            if pos >= len(buf):
                raise RedisRESPError("empty")
            prefix = chr(buf[pos])
            pos += 1

            if prefix == "+":
                line, pos2 = self._read_line(buf, pos)
                pos = pos2
                return line.decode()

            if prefix == "-":
                line, pos2 = self._read_line(buf, pos)
                pos = pos2
                raise RedisRESPError(line.decode())

            if prefix == ":":
                line, pos2 = self._read_line(buf, pos)
                pos = pos2
                return int(line.decode())

            if prefix == "$":
                line, pos2 = self._read_line(buf, pos)
                pos = pos2
                n = int(line.decode())
                if n == -1:
                    return None
                res = bytes(buf[pos : pos + n])
                pos += n + 2
                return res.decode(errors="replace")

            if prefix == "*":
                line, pos2 = self._read_line(buf, pos)
                pos = pos2
                n = int(line.decode())
                if n == -1:
                    return None
                arr = []
                for _ in range(n):
                    arr.append(parse_at())
                return arr

            raise RedisRESPError("unknown prefix")

        return parse_at()

    async def _send_raw(self, data: bytes):
        if self._ws is None or getattr(self._ws, "closed", False):
            await self.connect()
        await self._ws.send(data)

    async def _recv_raw(self) -> bytes:
        if self._ws is None or getattr(self._ws, "closed", False):
            await self.connect()
        msg = await self._ws.recv()
        return msg if isinstance(msg, (bytes, bytearray)) else msg.encode()

    async def execute(self, *parts: Union[str, int, float, bytes]) -> Any:
        async with self._lock:
            payload = self._encode_cmd(*parts)
            try:
                await self._send_raw(payload)
                data = await self._recv_raw()
            except Exception:
                await self.aclose()
                raise

            if not data or data[:1] not in b"+-:$*":
                await self.aclose()
                return None

            try:
                return self._parse_resp(data)
            except RedisRESPError:
                await self.aclose()
                return None

    # ===== Comandos =====
    async def ping(self) -> bool:
        r = await self.execute("PING")
        return (r == "PONG") or (r is not None)

    async def get(self, key: str) -> Optional[str]:
        return await self.execute("GET", key)

    async def set(self, key: str, value: Union[str, bytes], ex: Optional[int] = None):
        if ex is None:
            return await self.execute("SET", key, value)
        return await self.execute("SET", key, value, "EX", int(ex))

    async def delete(self, *keys: str):
        return await self.execute("DEL", *keys)

    async def sadd(self, key: str, *members: str):
        return await self.execute("SADD", key, *members)

    async def srem(self, key: str, *members: str):
        return await self.execute("SREM", key, *members)

    async def smembers(self, key: str) -> List[str]:
        r = await self.execute("SMEMBERS", key)
        return [] if not r else [str(x) for x in r]

    async def rpush(self, key: str, *values: str):
        return await self.execute("RPUSH", key, *values)

    async def lpush(self, key: str, *values: str):
        return await self.execute("LPUSH", key, *values)

    async def llen(self, key: str) -> int:
        r = await self.execute("LLEN", key)
        return int(r or 0)

    async def lrange(self, key: str, start: int, stop: int) -> List[str]:
        r = await self.execute("LRANGE", key, start, stop)
        return [] if not r else [str(x) for x in r]

    async def lpop(self, key: str) -> Optional[str]:
        r = await self.execute("LPOP", key)
        return None if r is None else str(r)

    async def lindex(self, key: str, index: int) -> Optional[str]:
        r = await self.execute("LINDEX", key, index)
        return None if r is None else str(r)

    async def scan(self, cursor: str = "0", match: Optional[str] = None, count: int = 10):
        cmd = ["SCAN", cursor]
        if match:
            cmd += ["MATCH", match]
        if count:
            cmd += ["COUNT", int(count)]
        r = await self.execute(*cmd)
        if not r:
            return "0", []
        cur = str(r[0])
        keys = [str(x) for x in r[1]] if len(r) > 1 and isinstance(r[1], list) else []
        return cur, keys

    async def ttl(self, key: str) -> Optional[int]:
        r = await self.execute("TTL", key)
        try:
            return int(r)
        except Exception:
            return None

    async def hmset(self, key: str, mapping: dict):
        parts = ["HMSET", key]
        for f, v in mapping.items():
            parts.extend([str(f), str(v)])
        return await self.execute(*parts)

    async def hget(self, key: str, field: str):
        return await self.execute("HGET", key, field)

    async def hgetall(self, key: str):
        r = await self.execute("HGETALL", key)
        if not r:
            return {}
        it = iter(r)
        out = {}
        for f, v in zip(it, it):
            out[str(f)] = str(v)
        return out

    async def hdel(self, key: str, *fields: str):
        return await self.execute("HDEL", key, *fields)

    async def zadd(self, key: str, score: float, member: str):
        return await self.execute("ZADD", key, score, member)

    async def zrem(self, key: str, *members: str):
        return await self.execute("ZREM", key, *members)

    async def zrange(self, key: str, start: int, stop: int, withscores=False):
        if withscores:
            r = await self.execute("ZRANGE", key, start, stop, "WITHSCORES")
            out = []
            it = iter(r)
            for m, s in zip(it, it):
                out.append((str(m), float(s)))
            return out
        else:
            r = await self.execute("ZRANGE", key, start, stop)
            return [] if not r else [str(x) for x in r]

    async def zrangebyscore(self, key: str, min_v: str, max_v: str, withscores=False):
        if withscores:
            r = await self.execute("ZRANGEBYSCORE", key, min_v, max_v, "WITHSCORES")
            out = []
            it = iter(r)
            for m, s in zip(it, it):
                out.append((str(m), float(s)))
            return out
        else:
            r = await self.execute("ZRANGEBYSCORE", key, min_v, max_v)
            return [] if not r else [str(x) for x in r]

    async def publish(self, channel: str, message: str):
        return await self.execute("PUBLISH", channel, message)


class AsyncRedisPubSubWS:
    """
    Conexão dedicada a Pub/Sub via RESP-over-WS.
    """

    def __init__(self, url: str, token: str = ""):
        self.url = f"{url}?token={token}" if token else url
        self._ws: Optional[websockets.WebSocketClientProtocol] = None

    async def connect(self):
        if self._ws and not getattr(self._ws, "closed", False):
            return
        dbg("[redis-pubsub] connecting ->", self.url)
        self._ws = await websockets.connect(self.url, max_size=None)
        dbg("[redis-pubsub] connected")

    async def aclose(self):
        if self._ws is not None:
            try:
                await self._ws.close()
                dbg("[redis-pubsub] closed")
            except Exception as e:
                dbg("[redis-pubsub] close error:", repr(e))
            self._ws = None

    def _enc(self, *parts: Union[str, bytes, int, float]) -> bytes:
        out = bytearray()
        out.extend(f"*{len(parts)}\r\n".encode())
        for p in parts:
            s = p if isinstance(p, (bytes, bytearray)) else str(p).encode()
            out.extend(f"${len(s)}\r\n".encode())
            out.extend(s)
            out.extend(b"\r\n")
        return bytes(out)

    async def _send(self, *parts):
        if not self._ws or getattr(self._ws, "closed", False):
            await self.connect()
        await self._ws.send(self._enc(*parts))

    async def subscribe(self, *channels: str):
        if not channels:
            return
        dbg("[redis-pubsub] SUBSCRIBE", channels)
        await self._send("SUBSCRIBE", *channels)

    async def psubscribe(self, *patterns: str):
        if not patterns:
            return
        dbg("[redis-pubsub] PSUBSCRIBE", patterns)
        await self._send("PSUBSCRIBE", *patterns)

    async def read_message(self) -> Optional[Dict[str, Any]]:
        """
        Retorna:
          {"type":"message","channel": "...", "payload": "..."}
          {"type":"pmessage","pattern":"...", "channel": "...", "payload": "..."}
          {"type":"subscribe"|...}
        """
        if not self._ws or getattr(self._ws, "closed", False):
            await self.connect()
        data = await self._ws.recv()
        if not isinstance(data, (bytes, bytearray)):
            data = str(data).encode()
        text = data.decode(errors="replace")
        parts = text.split("\r\n")
        if not parts or not parts[0].startswith("*"):
            return None
        try:
            elems = []
            i = 1
            while i < len(parts):
                line = parts[i]
                if line.startswith("$"):
                    i += 1
                    if i < len(parts):
                        elems.append(parts[i])
                i += 1
            if not elems:
                return None
            kind = elems[0]
            if kind == "message" and len(elems) >= 3:
                return {"type": "message", "channel": elems[1], "payload": elems[2]}
            if kind == "pmessage" and len(elems) >= 4:
                return {
                    "type": "pmessage",
                    "pattern": elems[1],
                    "channel": elems[2],
                    "payload": elems[3],
                }
            return {"type": kind, "elems": elems}
        except Exception:
            return None


# ===================== CONFIG =====================

REDIS_WS_URL = os.getenv("REDIS_WS_URL", "wss://redisrender.onrender.com")
REDIS_TOKEN = os.getenv("REDIS_TOKEN", "")

REDIS_LOG_NS = os.getenv("REDIS_LOG_NS", "iss")
PREQUEUE_NS = os.getenv("PREQUEUE_NS", f"{REDIS_LOG_NS}:y")
PREQUEUE_COLABS_SET = f"{PREQUEUE_NS}:colabs"

FRONT_LAST_RUN_KEY = f"{REDIS_LOG_NS}:front:last_run_ts"
ARQ_ENABLED = False  # informativo

SESSION_COOKIE_NAME = os.getenv("FRONT_SESSION_COOKIE", "iss_front_session")
SESSION_TTL = int(os.getenv("FRONT_SESSION_TTL", "86400"))  # 1 dia

# ===================== HELPERS =====================


def runs_key_for(tenant: str) -> str:
    return f"{REDIS_LOG_NS}:front:{tenant}:runs"


def session_key(token: str) -> str:
    return f"{REDIS_LOG_NS}:front:sess:{token}"


def cnpj_digits(cnpj: str) -> str:
    d = "".join(filter(str.isdigit, cnpj))[-14:]
    return d.zfill(14) if d else "0" * 14


def mask_cnpj(cnpj: str) -> str:
    d = cnpj_digits(cnpj)
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def slugify(value: str) -> str:
    import re
    import unicodedata

    value = (value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_") or "colab"


def task_logs_key(job_id: str) -> str:
    return f"{REDIS_LOG_NS}:task:{job_id}:logs"


def extract_meta(function: str, args: List[Any], kwargs: Dict[str, Any]) -> Dict[str, Any]:
    meta = {
        "cnpj": "",
        "colaborador": "",
        "colaborador_norm": "",
        "mes": "",
        "etapa": function or "",
    }
    try:
        if function == "job_notas" and len(args) >= 4:
            colab_norm, cnpj, mes = str(args[0]), str(args[1]), str(args[2])
            meta.update(
                {
                    "colaborador_norm": colab_norm,
                    "colaborador": colab_norm,
                    "cnpj": cnpj,
                    "mes": mes,
                    "etapa": "notas",
                }
            )
        elif function == "job_escrituracao" and len(args) >= 4:
            colab_norm, cnpj, mes = str(args[0]), str(args[1]), str(args[2])
            meta.update(
                {
                    "colaborador_norm": colab_norm,
                    "colaborador": colab_norm,
                    "cnpj": cnpj,
                    "mes": mes,
                    "etapa": "escrituracao",
                }
            )
        elif function == "job_certidao" and len(args) >= 4:
            colab_norm, cnpj, mes = str(args[0]), str(args[1]), str(args[2])
            meta.update(
                {
                    "colaborador_norm": colab_norm,
                    "colaborador": colab_norm,
                    "cnpj": cnpj,
                    "mes": mes,
                    "etapa": "certidao",
                }
            )
        elif function == "job_dam" and len(args) >= 3:
            colab_norm, cnpj = str(args[0]), str(args[1])
            meta.update(
                {
                    "colaborador_norm": colab_norm,
                    "colaborador": colab_norm,
                    "cnpj": cnpj,
                    "etapa": "dam",
                }
            )
    except Exception:
        pass
    return meta


def detect_etapa_from_logs(lines: List[str]) -> str:
    for line in lines:
        s = (line or "").lower()
        if "job_escrituracao" in s:
            return "escrituracao"
        if "job_notas" in s:
            return "notas"
        if "job_certidao" in s:
            return "certidao"
        if "job_dam" in s:
            return "dam"
    return ""


def detect_job_status_from_logs(lines: List[str]) -> str:
    if not lines:
        return "pending"
    lower_lines = [(l or "").lower() for l in lines]
    joined = "\n".join(lower_lines)
    if any("=== fim" in l and "ok" in l for l in lower_lines):
        return "success"
    if "'status': 'ok'" in joined or '"status": "ok"' in joined:
        return "success"
    if "finalizado" in joined or "concluído" in joined or "concluido" in joined:
        return "success"
    ERROR_TOKENS = [
        " erro ",
        "error",
        "exception",
        "traceback",
        "loginerror",
        "cnpjinexistenteerror",
        "cnpjmismatcherror",
    ]
    if any(tok in joined for tok in ERROR_TOKENS):
        return "error"
    return "running"


STATUS_PRIORITY = {"pending": 0, "waiting": 1, "queued": 2, "running": 3, "success": 4, "error": 4}


def merge_status(current: str, new: str) -> str:
    return new if STATUS_PRIORITY.get(new, 0) >= STATUS_PRIORITY.get(current, 0) else current


# ===================== APP =====================

app = FastAPI(title="ISS Queue Front (Cliente) — WS-only (Render)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def rds() -> AsyncRedisWS:
    # cliente de comandos (REST, leituras, gravações)
    return getattr(app.state, "redis_cmd", None)


async def get_session_from_request(request: Request) -> Optional[Dict[str, Any]]:
    rd: Optional[AsyncRedisWS] = getattr(app.state, "redis_cmd", None)
    if rd is None:
        dbg("get_session_from_request: redis_cmd not ready")
        return None
    token = request.cookies.get(SESSION_COOKIE_NAME)
    if not token:
        return None
    raw = await rd.get(session_key(token))
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


# ===================== AUTH MIDDLEWARE =====================

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in ("/login", "/logout", "/") or path.startswith("/openapi") or path.startswith("/docs") or path.startswith("/redoc"):
        return await call_next(request)

    if path == "/" or path.startswith("/api") or path.startswith("/ws"):
        sess = await get_session_from_request(request)
        if not sess:
            if path.startswith("/api"):
                dbg("auth_middleware: 401 on", path)
                return JSONResponse({"detail": "Não autenticado"}, status_code=401)
        else:
            request.state.session = sess
    return await call_next(request)


# ===================== STARTUP / SHUTDOWN =====================

@app.on_event("startup")
async def startup():
    # 1) comandos
    app.state.redis_cmd = AsyncRedisWS(REDIS_WS_URL, token=REDIS_TOKEN, name="cmd")
    await app.state.redis_cmd.connect()

    # 2) scheduler (isolado)
    app.state.redis_scheduler = AsyncRedisWS(REDIS_WS_URL, token=REDIS_TOKEN, name="sched")
    await app.state.redis_scheduler.connect()

    dbg("[startup] Redis WS OK (cmd/sched)")
    asyncio.create_task(scheduler_y_to_arq())


@app.on_event("shutdown")
async def shutdown():
    try:
        await app.state.redis_cmd.aclose()
    except Exception:
        pass
    try:
        await app.state.redis_scheduler.aclose()
    except Exception:
        pass
    print("Shutdown concluído (WS fechados).")


# ===================== RUNS PERSISTÊNCIA =====================

async def load_runs(rd: AsyncRedisWS, tenant: str) -> Dict[str, Dict[str, Any]]:
    if not tenant:
        return {}
    raw = await rd.get(runs_key_for(tenant))
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}


async def save_runs(rd: AsyncRedisWS, tenant: str, data: Dict[str, Dict[str, Any]]):
    if not tenant:
        return
    await rd.set(runs_key_for(tenant), json.dumps(data))


# ===================== HTML CLIENT + LOGIN =====================

PAGE_CLIENT = """<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8"/>
  <title>ISS Automation Client</title>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
</head>
<body>
  <!-- Aqui você pode injetar seu HTML/JS completo do front -->
  REPLACE_WITH_YOUR_FULL_PAGE_CLIENT_HTML
</body>
</html>
"""

PAGE_LOGIN = """<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8"/>
  <title>Login - ISS Automation</title>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <style>
    body { font-family: system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
           background: #0f172a; color: #e5e7eb; display: flex; align-items: center;
           justify-content: center; height: 100vh; margin: 0; }
    .card { background: #020617; padding: 24px 28px; border-radius: 16px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.4); width: 100%; max-width: 360px; }
    h1 { margin: 0 0 16px; font-size: 20px; }
    label { display: block; font-size: 14px; margin-bottom: 4px; }
    input { width: 100%; padding: 8px 10px; border-radius: 8px; border: 1px solid #1f2937;
            background: #020617; color: #e5e7eb; box-sizing: border-box; margin-bottom: 12px; }
    button { width: 100%; padding: 10px; border-radius: 8px; border: 0;
             background: #2563eb; color: white; font-weight: 600; cursor: pointer; }
    button:disabled { opacity: 0.6; cursor: default; }
    .error { color: #f97316; font-size: 13px; min-height: 18px; margin-top: 4px; }
  </style>
</head>
<body>
  <div class="card">
    <h1>Login ISS</h1>
    <form id="loginForm">
      <label for="colaborador">Usuário</label>
      <input id="colaborador" name="colaborador" autocomplete="username" required />

      <label for="senha">Senha</label>
      <input id="senha" name="senha" type="password" autocomplete="current-password" required />

      <button type="submit" id="btnLogin">Entrar</button>
      <div class="error" id="error"></div>
    </form>
  </div>
  <script>
    const form = document.getElementById("loginForm");
    const btn = document.getElementById("btnLogin");
    const err = document.getElementById("error");

    form.addEventListener("submit", async (e) => {
      e.preventDefault();
      err.textContent = "";
      btn.disabled = true;
      const fd = new FormData(form);
      try {
        const resp = await fetch("/login", {
          method: "POST",
          body: fd
        });
        if (!resp.ok) {
          const txt = await resp.text();
          err.textContent = txt || "Falha no login";
        } else {
          // Depois de logar, manda para a raiz (SPA)
          window.location.href = "/";
        }
      } catch (e) {
        err.textContent = "Erro de conexão";
      } finally {
        btn.disabled = false;
      }
    });
  </script>
</body>
</html>
"""


# ===================== ROTAS (REST) =====================

@app.get("/")
async def root():
    # Se quiser injetar um HTML gigante do front, substitua o placeholder
    html = PAGE_CLIENT.replace("REPLACE_WITH_YOUR_FULL_PAGE_CLIENT_HTML", "")
    return HTMLResponse(html, media_type="text/html")


@app.get("/login")
async def login_page():
    return HTMLResponse(PAGE_LOGIN, media_type="text/html")


@app.post("/login")
async def do_login(request: Request):
    rd = app.state.redis_cmd
    form = await request.form()
    colaborador = (form.get("colaborador") or "").strip().lower()
    senha = form.get("senha") or ""

    validUsers = {
        "valdinar": "123456",
        "julio": "123456",
        "samia": "123456",
        "alan": "123456",
        "isack": "123456",
        "thais": "123456",
        "legiscontabilidade": "iss2025@",
        "procontabil": "iss@2025",
        "laryssa": "123456",
        "leticiapro": "@iss2025",
        "alessandrapro": "@iss2025",
        "eucienepro": "@iss2025",
        "beatrizpro": "@iss2025",
        "luanpro": "@iss2025",
    }
    if colaborador not in validUsers or validUsers[colaborador] != senha:
        dbg("[login] inválido para", colaborador)
        raise HTTPException(401, "Usuário/senha inválidos")

    token = uuid.uuid4().hex
    sess_data = {
        "colaborador": colaborador,
        "colaborador_norm": slugify(colaborador),
        "created_at": time.time(),
    }
    await rd.set(session_key(token), json.dumps(sess_data), ex=SESSION_TTL)

    dbg("[login] ok para", colaborador, "token", token[:8] + "…")

    resp = JSONResponse({"status": "ok", "colaborador": colaborador})
    resp.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=token,
        max_age=SESSION_TTL,
        httponly=True,
        samesite="lax",
    )
    return resp


@app.get("/logout")
async def logout(request: Request):
    rd = app.state.redis_cmd
    token = request.cookies.get(SESSION_COOKIE_NAME)

    if token:
        try:
            await rd.delete(session_key(token))
            dbg("[logout] sessão removida", token[:8] + "…")
        except Exception as e:
            dbg("[logout] erro ao deletar sessão:", repr(e))

    resp = RedirectResponse(url="/login", status_code=302)
    resp.delete_cookie(SESSION_COOKIE_NAME)
    return resp


@app.get("/api/me")
async def api_me(request: Request):
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    return {
        "colaborador": sess.get("colaborador", ""),
        "colaborador_norm": sess.get("colaborador_norm", ""),
    }


import unicodedata


def normalize(s: str) -> str:
    s = s.strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace("_", " ").replace("-", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


@app.post("/api/import_cnpjs")
async def api_import_cnpjs(account_id: str = Form(...), file: UploadFile = File(...)):
    dbg("[import] recebendo arquivo:", file.filename, "para account:", account_id)

    from openpyxl import load_workbook
    from io import BytesIO

    def norm(s: str) -> str:
        s = str(s or "").strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        for ch in ["_", "-", ".", "/", "\\"]:
            s = s.replace(ch, " ")
        while "  " in s:
            s = s.replace("  ", " ")
        return s

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    headers_raw = []
    headers_norm = []
    rows_out = []

    for row_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_i == 1:
            headers_raw = [str(x or "") for x in row]
            headers_norm = [norm(x) for x in headers_raw]
            dbg(">>>> HEADERS ORIGINAIS:", headers_raw)
            dbg(">>>> HEADERS NORMALIZADOS:", headers_norm)
            continue

        record = {}
        for col_i, cell in enumerate(row):
            key = headers_norm[col_i] if col_i < len(headers_norm) else f"col{col_i}"
            record[key] = cell

        dbg(f"[import] Linha {row_i} record=", record)

        cnpj = str(record.get("cnpj") or "").strip()

        razao = str(
            record.get("razao social")
            or record.get("empresa")
            or record.get("razao")
            or record.get("nome")
            or ""
        ).strip()

        dominio = str(
            record.get("codigo dominio")
            or record.get("codigo")
            or record.get("codico")
            or record.get("dominio")
            or ""
        ).strip()

        dbg(f"[import] Linha {row_i}: cnpj='{cnpj}' razao='{razao}' dominio='{dominio}'")

        if not cnpj:
            continue

        rows_out.append({"cnpj": cnpj, "razao": razao, "dominio": dominio})

    dbg(f"[import] TOTAL extraídos: {len(rows_out)} registros")
    return {"cnpjs": rows_out}


@app.post("/api/enqueue")
async def api_enqueue(request: Request, payload: Dict[str, Any]):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm")
    if not tenant:
        raise HTTPException(401, "Sessão inválida")

    runs = await load_runs(rd, tenant)

    provider = str(payload.get("provider") or "1")
    mes = str(payload.get("mes") or "").zfill(2)
    ano = str(payload.get("ano") or "")
    etapas = payload.get("etapas") or {}
    formato_dominio = bool(payload.get("formato_dominio"))
    cnpj_ids = payload.get("cnpj_ids") or []

    accounts_list = payload.get("accounts") or []
    cnpjs_list = payload.get("cnpjs") or []
    accs = {str(a.get("id")): a for a in accounts_list if a.get("id")}
    cnps = {str(c.get("id")): c for c in cnpjs_list if c.get("id")}

    if not cnpj_ids:
        raise HTTPException(400, "Nenhum CNPJ selecionado.")

    mes_str = f"{mes}/{ano}"
    tipo_estrutura = "dominio" if formato_dominio else "convencional"

    func_map = {
        "escrituracao": "job_escrituracao",
        "notas": "job_notas",
        "dam": "job_dam",
        "certidao": "job_certidao",
    }

    total_jobs = 0
    used_cnpjs = set()

    for cid in cnpj_ids:
        c = cnps.get(cid)
        if not c:
            continue

        account_id = c.get("account_id") or ""
        acc = accs.get(account_id)
        if not acc:
            continue

        usuario = (acc.get("user") or "").strip()
        senha = acc.get("password") or ""
        if not usuario or not senha:
            continue

        colaborador_norm = tenant
        cd = cnpj_digits(str(c.get("cnpj") or ""))
        if not cd or cd == "0" * 14:
            continue

        used_cnpjs.add(cd)

        run_id = uuid.uuid4().hex
        run_job_ids: List[str] = []
        run_etapas_flags = {
            "escrituracao": False,
            "notas": False,
            "dam": False,
            "certidao": False,
        }

        codigo_dom = (c.get("dominio") or "").strip() or None

        for etapa_key, enabled in etapas.items():
            if not enabled:
                continue
            func = func_map.get(etapa_key)
            if not func:
                continue

            if func == "job_notas":
                args = [
                    colaborador_norm,
                    cd,
                    mes_str,
                    usuario,
                    senha,
                    tipo_estrutura,
                    codigo_dom,
                ]
            else:
                args = [colaborador_norm, cd, mes_str, usuario, senha, tipo_estrutura]

            job_id = uuid.uuid4().hex
            env = {
                "func": func,
                "args": args,
                "enqueue_kwargs": {"_job_id": job_id},
                "colaborador_norm": colaborador_norm,
            }

            y_key = f"{PREQUEUE_NS}:colab:{colaborador_norm}"
            await rd.sadd(PREQUEUE_COLABS_SET, colaborador_norm)
            await rd.rpush(y_key, json.dumps(env))
            total_jobs += 1
            run_job_ids.append(job_id)
            run_etapas_flags[etapa_key] = True

        if run_job_ids:
            runs[run_id] = {
                "id": run_id,
                "cnpj": cd,
                "mes": mes,
                "ano": ano,
                "mes_str": mes_str,
                "etapas": run_etapas_flags,
                "job_ids": run_job_ids,
                "account_id": account_id,
                "provider": provider,
                "tipo_estrutura": tipo_estrutura,
                "colaborador_norm": colaborador_norm,
                "created_at": time.time(),
            }

    if total_jobs == 0:
        raise HTTPException(400, "Nenhum job foi enfileirado (verifique senhas e contas).")

    await save_runs(rd, tenant, runs)
    await rd.publish(f"{REDIS_LOG_NS}:front:{tenant}:runs:updated", json.dumps(runs))
    await rd.set(FRONT_LAST_RUN_KEY, str(time.time()))

    dbg(f"[enqueue] tenant={tenant} jobs={total_jobs} cnpjs={len(used_cnpjs)} mes={mes_str}")
    return {
        "status": "ok",
        "total_jobs": total_jobs,
        "total_cnpjs": len(used_cnpjs),
        "arq_enabled": ARQ_ENABLED,
    }


@app.get("/api/kpis")
async def api_kpis(request: Request):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")

    runs = await load_runs(rd, tenant)

    rel_cnpjs: Dict[str, Dict[str, Any]] = {}
    distinct_cnpjs = set()
    for run in runs.values():
        cd = cnpj_digits(run.get("cnpj", ""))
        mes_str = run.get("mes_str") or ""
        if not cd:
            continue
        distinct_cnpjs.add(cd)
        key = f"{cd}:{mes_str}" if mes_str else cd
        rel_cnpjs[key] = {"cnpj": cd, "mes": mes_str}

    idx = await build_status_index(rel_cnpjs, allowed_colab_prefix=tenant)
    kpis = idx["kpis"]
    kpis["total_cnpjs"] = len(distinct_cnpjs)

    last_raw = await rd.get(FRONT_LAST_RUN_KEY)
    last_str = "—"
    if last_raw:
        try:
            ts = float(last_raw)
            last_str = time.strftime("%d/%m %H:%M", time.localtime(ts))
        except Exception:
            pass
    kpis["last_run_str"] = last_str
    kpis["arq_enabled"] = ARQ_ENABLED
    return {"kpis": kpis}


# ===================== STATUS INDEX =====================

async def build_status_index(
    relevant_cnpjs: Dict[str, Dict[str, Any]],
    allowed_colab_prefix: Optional[str] = None,
) -> Dict[str, Any]:
    rd: AsyncRedisWS = rds()
    job_events: Dict[str, Dict[str, Any]] = {}
    etapas = ["escrituracao", "notas", "dam", "certidao"]

    # --- normaliza entrada inicial (pode vir vazia) ---
    keys = list(relevant_cnpjs.keys())  # existing keys coming from runs

    # ---- varredura da pré-fila para acrescentar CNPJs que só estão em waiting ----
    try:
        colabs = await rd.smembers(PREQUEUE_COLABS_SET)
    except Exception:
        colabs = []

    for colab in colabs:
        y_key = f"{PREQUEUE_NS}:colab:{colab}"
        try:
            raw_items = await rd.lrange(y_key, 0, -1)
        except Exception:
            raw_items = []
        for raw in raw_items:
            try:
                env = json.loads(raw)
            except Exception:
                continue
            func = env.get("func") or ""
            args = env.get("args") or []
            meta = extract_meta(func, args, {})
            cd = cnpj_digits(meta.get("cnpj", ""))
            mes = meta.get("mes", "") or ""
            key = f"{cd}:{mes}" if mes else cd
            if cd and key not in relevant_cnpjs:
                relevant_cnpjs[key] = {"cnpj": cd, "mes": mes}
                keys.append(key)
    # ---- fim inclusão da pré-fila ----

    if not keys:
        return {
            "status_per_cnpj": {},
            "job_events": {},
            "kpis": {"running": 0, "queue": 0, "total_cnpjs": 0},
        }

    # inicializa estrutura events
    events: Dict[str, Dict[str, List[Dict[str, Any]]]] = {k: {e: [] for e in etapas} for k in keys}

    running_count = 0
    queue_count = 0

    def add_job_event(job_id: Optional[str], cd: str, mes: str, etapa: str, status: str):
        if not job_id:
            return
        prev = job_events.get(job_id)
        if not prev:
            job_events[job_id] = {
                "cnpj": cd,
                "mes": mes or "",
                "etapa": etapa,
                "status": status,
            }
        else:
            prev_status = prev.get("status", "pending")
            if STATUS_PRIORITY.get(status, 0) >= STATUS_PRIORITY.get(prev_status, 0):
                prev.update(
                    {
                        "status": status,
                        "cnpj": cd or prev.get("cnpj", ""),
                        "mes": mes or prev.get("mes", ""),
                        "etapa": etapa or prev.get("etapa", ""),
                    }
                )

    def append_to_event_maps(cd: str, mes: str, etapa: str, st_str: str, job_id: str):
        key1 = cd
        key2 = f"{cd}:{mes}" if mes else None
        if key2 and key2 in events:
            events[key2].setdefault(etapa, []).append({"status": st_str, "job_id": job_id})
        if key1 in events:
            events[key1].setdefault(etapa, []).append({"status": st_str, "job_id": job_id})

    # 1) Pré-fila (waiting)
    for colab in colabs:
        y_key = f"{PREQUEUE_NS}:colab:{colab}"
        try:
            raw_items = await rd.lrange(y_key, 0, -1)
        except Exception:
            raw_items = []
        for raw in raw_items:
            try:
                env = json.loads(raw)
            except Exception:
                continue
            func = env.get("func") or ""
            args = env.get("args") or []
            meta = extract_meta(func, args, {})
            cd = cnpj_digits(meta["cnpj"])
            mes = meta.get("mes", "") or ""
            etapa = meta["etapa"] or func
            colab_norm = meta.get("colaborador_norm") or ""
            if allowed_colab_prefix and not str(colab_norm).startswith(allowed_colab_prefix):
                continue
            jid = env.get("enqueue_kwargs", {}).get("_job_id")
            st_str = "waiting"
            append_to_event_maps(cd, mes, etapa, st_str, jid or "")
            add_job_event(jid, cd, mes, etapa, st_str)
            queue_count += 1

    # 2) Jobs já rodados (logs das runs)
    tenant = allowed_colab_prefix or ""
    runs_for_tenant: Dict[str, Any] = {}
    if tenant:
        try:
            runs_for_tenant = await load_runs(rd, tenant)
        except Exception:
            runs_for_tenant = {}

    for run in runs_for_tenant.values():
        cd = cnpj_digits(run.get("cnpj", ""))
        if not cd:
            continue

        mes_str = run.get("mes_str") or ""
        if not mes_str:
            mes_only = str(run.get("mes") or "").zfill(2)
            ano_only = str(run.get("ano") or "").strip()
            mes_str = f"{mes_only}/{ano_only}" if mes_only and ano_only else mes_only

        job_ids = run.get("job_ids") or []
        for jid in job_ids:
            try:
                jlogs = await rd.lrange(task_logs_key(jid), 0, -1)
            except Exception:
                jlogs = []
            if not jlogs:
                continue

            etapa = detect_etapa_from_logs(jlogs) or ""
            if not etapa:
                flags = run.get("etapas") or {}
                true_steps = [k for k, v in flags.items() if v]
                if len(true_steps) == 1:
                    etapa = true_steps[0]

            status = detect_job_status_from_logs(jlogs)
            if not etapa:
                add_job_event(jid, cd, mes_str, "", status)
                continue

            append_to_event_maps(cd, mes_str, etapa, status, jid)
            add_job_event(jid, cd, mes_str, etapa, status)
            if status == "running":
                running_count += 1

    def reduce_status(lst: List[Dict[str, Any]]) -> str:
        if not lst:
            return "pending"
        flags = [e["status"] for e in lst]
        if "running" in flags:
            return "running"
        if "success" in flags and "error" not in flags:
            return "success"
        if "error" in flags:
            return "error"
        if "waiting" in flags or "queued" in flags:
            return "waiting"
        return "pending"

    status_per_cnpj: Dict[str, Dict[str, str]] = {}
    for key in events:
        s = {etapa: reduce_status(events[key].get(etapa, [])) for etapa in etapas}
        status_per_cnpj[key] = s

    return {
        "status_per_cnpj": status_per_cnpj,
        "job_events": job_events,
        "kpis": {
            "running": running_count,
            "queue": queue_count,
            "total_cnpjs": len(keys),
        },
    }


@app.get("/api/status")
async def api_status(request: Request):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")

    runs = await load_runs(rd, tenant)
    rel_cnpjs: Dict[str, Dict[str, Any]] = {}
    for run in runs.values():
        cd = cnpj_digits(run.get("cnpj", ""))
        mes_str = run.get("mes_str") or ""
        if cd:
            key = f"{cd}:{mes_str}" if mes_str else cd
            rel_cnpjs[key] = {"cnpj": cd, "mes": mes_str}
    idx = await build_status_index(rel_cnpjs, allowed_colab_prefix=tenant)
    kpis = idx["kpis"]
    job_events: Dict[str, Dict[str, Any]] = idx.get("job_events", {})

    def reduce_status_list(statuses: List[str]) -> str:
        if not statuses:
            return "pending"
        cur = "pending"
        for st in statuses:
            cur = merge_status(cur, st)
        return cur

    rows: List[Dict[str, Any]] = []
    for run_id, run in runs.items():
        cd = cnpj_digits(run.get("cnpj", ""))
        if not cd:
            continue
        job_ids: List[str] = run.get("job_ids") or []
        etapas_status = {
            "escrituracao": "pending",
            "notas": "pending",
            "dam": "pending",
            "certidao": "pending",
        }
        etapa_to_statuses: Dict[str, List[str]] = {k: [] for k in etapas_status.keys()}

        for jid in job_ids:
            je = job_events.get(jid)
            if not je:
                continue
            etapa = (je.get("etapa") or "").lower()
            st = je.get("status") or "pending"
            if etapa in etapa_to_statuses:
                etapa_to_statuses[etapa].append(st)

        for e in etapa_to_statuses:
            etapas_status[e] = reduce_status_list(etapa_to_statuses[e])
        geral_status = reduce_status_list(list(etapas_status.values()))

        rows.append(
            {
                "run_id": run_id,
                "cnpj_digits": cd,
                "cnpj_mask": mask_cnpj(cd),
                "razao": "",
                "conta": "",
                "mes": run.get("mes") or "",
                "ano": run.get("ano") or "",
                "etapas": etapas_status,
                "status_geral": geral_status,
            }
        )
    rows.sort(key=lambda r: (r["cnpj_digits"], r["ano"], r["mes"], r["run_id"]))
    return {"rows": rows, "kpis": kpis, "arq_enabled": ARQ_ENABLED}


@app.get("/api/logs/{cnpj}")
async def api_logs_for_cnpj(request: Request, cnpj: str):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")
    colab_prefix = tenant + "::" if tenant else ""

    cn = cnpj_digits(cnpj.strip().replace(".", "").replace("/", "").replace("-", ""))
    mes = (request.query_params.get("mes") or "").zfill(2)
    ano = (request.query_params.get("ano") or "").strip()
    mes_key = f"{mes}/{ano}" if mes and ano else (mes if mes else "")
    key = f"{cn}:{mes_key}" if mes_key else cn
    rel_cnpjs = {key: {"cnpj": cn, "mes": mes_key}}
    idx = await build_status_index(rel_cnpjs, allowed_colab_prefix=tenant)
    etapas_status = idx["status_per_cnpj"].get(key) or idx["status_per_cnpj"].get(cn, {})

    logs_lines: List[str] = []
    job_ids_for_cnpj = set()

    runs = await load_runs(rd, tenant)
    for r in runs.values():
        if cnpj_digits(r.get("cnpj", "")) == cn:
            for jid in (r.get("job_ids") or []):
                job_ids_for_cnpj.add(jid)

    try:
        colabs = await rd.smembers(PREQUEUE_COLABS_SET)
    except Exception:
        colabs = []
    for colab in colabs:
        if colab_prefix and not str(colab).startswith(colab_prefix):
            continue
        y_key = f"{PREQUEUE_NS}:colab:{colab}"
        try:
            raw_items = await rd.lrange(y_key, 0, -1)
        except Exception:
            raw_items = []
        for raw_item in raw_items:
            try:
                env = json.loads(raw_item)
            except Exception:
                continue
            meta = extract_meta(env.get("func") or "", env.get("args") or [], {})
            colab_norm = meta.get("colaborador_norm") or ""
            if colab_prefix and not colab_norm.startswith(colab_prefix):
                continue
            if cnpj_digits(meta["cnpj"]) == cn:
                jid = env.get("enqueue_kwargs", {}).get("_job_id", "")
                if jid:
                    job_ids_for_cnpj.add(jid)

    for jid in sorted(job_ids_for_cnpj):
        try:
            jlogs = await rd.lrange(task_logs_key(jid), 0, -1)
        except Exception:
            jlogs = []
        if jlogs:
            logs_lines.append(f"=== Job {jid} ===")
            logs_lines.extend(jlogs)

    return {
        "cnpj": cn,
        "cnpj_mask": mask_cnpj(cn),
        "razao": "",
        "etapas": {
            "escrituracao": etapas_status.get("escrituracao", "pending"),
            "notas": etapas_status.get("notas", "pending"),
            "dam": etapas_status.get("dam", "pending"),
            "certidao": etapas_status.get("certidao", "pending"),
        },
        "logs": "\n".join(logs_lines),
        "arq_enabled": ARQ_ENABLED,
    }


@app.get("/api/logs_job/{job_id}")
async def api_logs_for_job(request: Request, job_id: str):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")
    colab_prefix = tenant + "::" if tenant else ""

    cn = ""
    etapa = ""
    status = "pending"
    mes = ""

    try:
        colabs = await rd.smembers(PREQUEUE_COLABS_SET)
    except Exception:
        colabs = []
    for colab in colabs:
        if colab_prefix and not str(colab).startswith(colab_prefix):
            continue
        y_key = f"{PREQUEUE_NS}:colab:{colab}"
        try:
            raw_items = await rd.lrange(y_key, 0, -1)
        except Exception:
            raw_items = []
        found = False
        for raw_item in raw_items:
            try:
                env = json.loads(raw_item)
            except Exception:
                continue
            jid = env.get("enqueue_kwargs", {}).get("_job_id")
            if jid != job_id:
                continue
            func = env.get("func") or ""
            args = env.get("args") or []
            meta = extract_meta(func, args, {})
            colab_norm = meta.get("colaborador_norm") or ""
            if colab_prefix and not colab_norm.startswith(colab_prefix):
                continue
            cn = cnpj_digits(meta["cnpj"])
            etapa = meta.get("etapa") or func
            mes = meta.get("mes") or ""
            status = "waiting"
            found = True
            break
        if found:
            break

    if not cn:
        runs = await load_runs(rd, tenant)
        for r in runs.values():
            if job_id in (r.get("job_ids") or []):
                cn = cnpj_digits(r.get("cnpj", ""))
                mes = r.get("mes_str") or ""
                if not mes:
                    mes_only = (r.get("mes") or "").zfill(2)
                    ano_only = str(r.get("ano") or "").strip()
                    mes = f"{mes_only}/{ano_only}" if mes_only and ano_only else (mes_only or "")
                break

    if not cn:
        raise HTTPException(404, "Job não encontrado para este usuário.")

    try:
        jlogs = await rd.lrange(task_logs_key(job_id), 0, -1)
    except Exception:
        jlogs = []
    logs = "\n".join(jlogs or [])

    if jlogs:
        status = detect_job_status_from_logs(jlogs)

    etapas_status: Dict[str, str] = {}
    if cn:
        key = f"{cn}:{mes}" if mes else cn
        rel_cnpjs = {key: {"cnpj": cn, "mes": mes}}
        idx = await build_status_index(rel_cnpjs, allowed_colab_prefix=tenant)
        etapas_status = idx["status_per_cnpj"].get(key) or idx["status_per_cnpj"].get(cn, {})

    return {
        "job_id": job_id,
        "cnpj": cn,
        "cnpj_mask": mask_cnpj(cn) if cn else "",
        "razao": "",
        "etapa": etapa,
        "status": status,
        "etapas": {
            "escrituracao": etapas_status.get("escrituracao", "pending"),
            "notas": etapas_status.get("notas", "pending"),
            "dam": etapas_status.get("dam", "pending"),
            "certidao": etapas_status.get("certidao", "pending"),
        },
        "logs": logs,
        "arq_enabled": ARQ_ENABLED,
    }


@app.get("/api/logs_run/{run_id}")
async def api_logs_for_run(request: Request, run_id: str):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")

    runs = await load_runs(rd, tenant)
    run = runs.get(run_id)
    if not run:
        raise HTTPException(404, "Execução não encontrada.")

    cn = cnpj_digits(run.get("cnpj", ""))
    job_ids: List[str] = run.get("job_ids") or []

    etapas_status_run: Dict[str, str] = {
        "escrituracao": "pending",
        "notas": "pending",
        "dam": "pending",
        "certidao": "pending",
    }
    logs_lines: List[str] = []

    for jid in job_ids:
        try:
            jlogs = await rd.lrange(task_logs_key(jid), 0, -1)
        except Exception:
            jlogs = []
        if not jlogs:
            continue
        logs_lines.append(f"=== Job {jid} ===")
        logs_lines.extend(jlogs)
        etapa_detectada = detect_etapa_from_logs(jlogs)
        job_status = detect_job_status_from_logs(jlogs)
        if etapa_detectada:
            cur = etapas_status_run.get(etapa_detectada, "pending")
            etapas_status_run[etapa_detectada] = merge_status(cur, job_status)

    return {
        "run_id": run_id,
        "cnpj": cn,
        "cnpj_mask": mask_cnpj(cn),
        "razao": "",
        "mes": run.get("mes") or "",
        "ano": run.get("ano") or "",
        "etapas": etapas_status_run,
        "logs": "\n".join(logs_lines),
        "arq_enabled": ARQ_ENABLED,
    }


@app.post("/api/stop_all")
async def api_stop_all(request: Request):
    rd = app.state.redis_cmd
    sess = getattr(request.state, "session", None) or await get_session_from_request(request)
    if not sess:
        raise HTTPException(401, "Não autenticado")
    tenant = sess.get("colaborador_norm", "")
    tenant_prefix = tenant + "::" if tenant else ""

    runs = await load_runs(rd, tenant)
    colabs = set()
    try:
        prequeue_colabs = await rd.smembers(PREQUEUE_COLABS_SET)
        for c in prequeue_colabs:
            if tenant_prefix and str(c).startswith(tenant_prefix):
                colabs.add(str(c))
    except Exception:
        pass
    for r in runs.values():
        colab_norm = r.get("colaborador_norm")
        if colab_norm and (not tenant_prefix or str(colab_norm).startswith(tenant_prefix)):
            colabs.add(str(colab_norm))

    prequeue_removed = 0

    for c in colabs:
        key = f"{PREQUEUE_NS}:colab:{c}"
        try:
            n = await rd.llen(key)
            if n and n > 0:
                await rd.delete(key)
                prequeue_removed += n
        except Exception:
            pass
        try:
            await rd.srem(PREQUEUE_COLABS_SET, c)
        except Exception:
            pass

    dbg(f"[stop_all] tenant={tenant} prequeue_removed={prequeue_removed}")
    return {
        "status": "ok",
        "stopped_jobs": 0,
        "prequeue_removed": prequeue_removed,
        "ids": [],
        "arq_enabled": ARQ_ENABLED,
    }


from pcloud_download import make_pcloud_router

app.include_router(make_pcloud_router(get_session_from_request))


# ===================== WEBSOCKET: PUSH runs/status/log =====================

class LiveHub:
    def __init__(self):
        self.tenants_clients: Dict[str, Set[WebSocket]] = {}
        self.tenants_sub_task: Dict[str, asyncio.Task] = {}
        self.lock = asyncio.Lock()

    async def attach(self, tenant: str, ws: WebSocket):
        async with self.lock:
            self.tenants_clients.setdefault(tenant, set()).add(ws)
            if tenant not in self.tenants_sub_task:
                self.tenants_sub_task[tenant] = asyncio.create_task(self._run_subscriber_loop(tenant))
        dbg("[livehub] attach tenant=", tenant, "clients=", len(self.tenants_clients.get(tenant, set())))

    async def detach(self, tenant: str, ws: WebSocket):
        async with self.lock:
            if tenant in self.tenants_clients:
                self.tenants_clients[tenant].discard(ws)
        dbg("[livehub] detach tenant=", tenant)

    async def _broadcast(self, tenant: str, payload: Dict[str, Any]):
        conns = self.tenants_clients.get(tenant, set())
        for c in list(conns):
            try:
                await c.send_json(payload)
            except Exception:
                conns.discard(c)
        dbg("[livehub] broadcast", payload.get("type"), "to", len(conns), "client(s)")

    async def _run_subscriber_loop(self, tenant: str):
        pubsub = AsyncRedisPubSubWS(REDIS_WS_URL, token=REDIS_TOKEN)
        await pubsub.connect()

        runs_channel = f"{REDIS_LOG_NS}:front:{tenant}:runs:updated"
        await pubsub.subscribe(runs_channel)

        known_job_channels: Set[str] = set()
        runs: Dict[str, Any] = {}

        # Snapshot inicial
        try:
            runs = await load_runs(app.state.redis_cmd, tenant)
            await self._broadcast(tenant, {"type": "runs", "runs": runs})
            rows, kpis = await compute_status_for_tenant(tenant, runs)
            await self._broadcast(tenant, {"type": "status", "rows": rows, "kpis": kpis})

            for r in runs.values():
                for jid in (r.get("job_ids") or []):
                    ch = f"{REDIS_LOG_NS}:task:{jid}:logs"
                    if ch not in known_job_channels:
                        await pubsub.subscribe(ch)
                        known_job_channels.add(ch)
        except Exception as e:
            print("[livehub] snapshot inicial erro:", repr(e))

        try:
            while True:
                try:
                    msg = await pubsub.read_message()
                    if not msg:
                        await asyncio.sleep(0.05)
                        continue

                    if msg["type"] == "message":
                        channel = msg["channel"]
                        payload = msg.get("payload") or ""

                        if channel == runs_channel:
                            try:
                                runs = json.loads(payload or "{}")
                            except Exception:
                                runs = {}
                            await self._broadcast(tenant, {"type": "runs", "runs": runs})

                            new_jobs: Set[str] = set()
                            for r in runs.values():
                                for jid in (r.get("job_ids") or []):
                                    new_jobs.add(jid)
                            for jid in new_jobs:
                                ch = f"{REDIS_LOG_NS}:task:{jid}:logs"
                                if ch not in known_job_channels:
                                    await pubsub.subscribe(ch)
                                    known_job_channels.add(ch)

                            asyncio.create_task(self._push_status_async(tenant, runs))

                        elif channel.startswith(f"{REDIS_LOG_NS}:task:") and channel.endswith(":logs"):
                            try:
                                jid = channel.split(":task:")[1].rsplit(":", 1)[0]
                            except Exception:
                                jid = ""
                            await self._broadcast(
                                tenant,
                                {
                                    "type": "log",
                                    "job_id": jid,
                                    "line": payload,
                                    "channel": channel,
                                },
                            )
                            asyncio.create_task(self._push_status_async(tenant, runs))
                except Exception as e:
                    print("[livehub] loop erro:", repr(e))
                    await asyncio.sleep(0.25)
        finally:
            await pubsub.aclose()

    async def _push_status_async(self, tenant: str, runs: Dict[str, Any]):
        try:
            rows, kpis = await compute_status_for_tenant(tenant, runs)
            await self._broadcast(tenant, {"type": "status", "rows": rows, "kpis": kpis})
        except Exception as e:
            print("[livehub] _push_status_async erro:", repr(e))


livehub = LiveHub()


async def compute_status_for_tenant(tenant: str, runs: Dict[str, Any]):
    rel_cnpjs: Dict[str, Dict[str, Any]] = {}
    for run in runs.values():
        cd = cnpj_digits(run.get("cnpj", ""))
        if not cd:
            continue
        mes_str = run.get("mes_str") or ""
        key = f"{cd}:{mes_str}" if mes_str else cd
        rel_cnpjs[key] = {"cnpj": cd, "mes": mes_str}

    idx = await build_status_index(rel_cnpjs, allowed_colab_prefix=tenant)
    kpis = idx["kpis"]
    job_events: Dict[str, Dict[str, Any]] = idx.get("job_events", {})

    def reduce_status_list(statuses: List[str]) -> str:
        if not statuses:
            return "pending"
        cur = "pending"
        for st in statuses:
            cur = merge_status(cur, st)
        return cur

    rows: List[Dict[str, Any]] = []
    for run_id, run in runs.items():
        cd = cnpj_digits(run.get("cnpj", ""))
        if not cd:
            continue
        job_ids: List[str] = run.get("job_ids") or []
        etapas_status = {
            "escrituracao": "pending",
            "notas": "pending",
            "dam": "pending",
            "certidao": "pending",
        }
        etapa_to_statuses: Dict[str, List[str]] = {k: [] for k in etapas_status.keys()}

        for jid in job_ids:
            je = job_events.get(jid)
            if not je:
                continue
            etapa = (je.get("etapa") or "").lower()
            st = je.get("status") or "pending"
            if etapa in etapa_to_statuses:
                etapa_to_statuses[etapa].append(st)

        for e in etapa_to_statuses:
            etapas_status[e] = reduce_status_list(etapa_to_statuses[e])
        geral_status = reduce_status_list(list(etapas_status.values()))

        rows.append(
            {
                "run_id": run_id,
                "cnpj_digits": cd,
                "cnpj_mask": mask_cnpj(cd),
                "razao": "",
                "conta": "",
                "mes": run.get("mes") or "",
                "ano": run.get("ano") or "",
                "etapas": etapas_status,
                "status_geral": geral_status,
            }
        )
    rows.sort(key=lambda r: (r["cnpj_digits"], r["ano"], r["mes"], r["run_id"]))

    last_raw = await app.state.redis_cmd.get(FRONT_LAST_RUN_KEY)
    last_str = "—"
    if last_raw:
        try:
            ts = float(last_raw)
            last_str = time.strftime("%d/%m %H:%M", time.localtime(ts))
        except Exception:
            pass
    kpis["last_run_str"] = last_str
    kpis["arq_enabled"] = ARQ_ENABLED

    return rows, kpis


@app.websocket("/ws")
async def ws_bridge(websocket: WebSocket):
    await websocket.accept()
    try:
        cookie_header = websocket.headers.get("cookie", "") or websocket.headers.get("Cookie", "")
        token = ""
        for part in cookie_header.split(";"):
            part = part.strip()
            if not part:
                continue
            if part.startswith(SESSION_COOKIE_NAME + "="):
                token = part.split("=", 1)[1]
                break

        if not token:
            await websocket.send_json({"type": "error", "detail": "Não autenticado"})
            await websocket.close()
            return

        rd = app.state.redis_cmd
        raw = await rd.get(session_key(token))
        if not raw:
            await websocket.send_json({"type": "error", "detail": "Sessão inválida"})
            await websocket.close()
            return

        sess = json.loads(raw)
        tenant = sess.get("colaborador_norm") or ""
        if not tenant:
            await websocket.send_json({"type": "error", "detail": "Sessão sem tenant"})
            await websocket.close()
            return
    except Exception as e:
        await websocket.send_json({"type": "error", "detail": f"Falha de sessão: {e!s}"})
        await websocket.close()
        return

    await websocket.send_json({"type": "hello", "tenant": tenant})
    await livehub.attach(tenant, websocket)

    try:
        while True:
            try:
                _ = await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
            except asyncio.TimeoutError:
                await websocket.send_json({"type": "ping", "ts": time.time()})
    except WebSocketDisconnect:
        pass
    finally:
        await livehub.detach(tenant, websocket)


# ===================== SCHEDULER INTERNO — Y ➜ ARQ =====================

async def scheduler_y_to_arq():
    print("[scheduler] iniciado!")
    rd: AsyncRedisWS = app.state.redis_scheduler

    while True:
        try:
            colabs = await rd.smembers(PREQUEUE_COLABS_SET)
            if not colabs:
                await asyncio.sleep(1)
                continue

            for colab in colabs:
                y_key = f"{PREQUEUE_NS}:colab:{colab}"

                raw = await rd.lpop(y_key)
                if not raw:
                    continue

                env = json.loads(raw)
                func = env.get("func")
                args_list = env.get("args") or []
                job_id = env.get("enqueue_kwargs", {}).get("_job_id")
                if not job_id:
                    continue

                now_ms = int(time.time() * 1000)

                jobdef = {"t": 1, "f": func, "a": tuple(args_list), "k": {}, "et": now_ms}

                await rd.set(f"arq:job:{job_id}", pickle.dumps(jobdef))
                await rd.zadd("arq:queue", now_ms, job_id)

                print(f"[scheduler] job {job_id} → ARQ OK")
        except Exception as e:
            print("[scheduler] erro:", repr(e))

        await asyncio.sleep(0.25)


# ===================== MAIN =====================

if __name__ == "__main__":
    import uvicorn

    uvicorn.run("queue_client_front:app", host="0.0.0.0", port=8001, reload=True)
